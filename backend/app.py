from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import mimetypes
from typing import Dict, List
import io
import re
import traceback
import csv
import requests
import json

# 파일 처리 라이브러리
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from PIL import Image
    from PIL.ExifTags import TAGS
except ImportError:
    Image = None
    TAGS = None

try:
    import docx
except ImportError:
    docx = None


app = Flask(__name__)
CORS(app)
app.config['DEBUG'] = True

# PII 패턴 정의 
PII_PATTERNS = {
    "KR_RRN": {
        "label": "주민등록번호",
        "regex": r"(?:[0-9]{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12][0-9]|3[01]))-[1-4][0-9]{6}",
        "score": 10
    },
    "PASSPORT": {
        "label": "여권번호",
        "regex": r"[MS][A-Z0-9]{8,9}",
        "score": 10
    },
    "DRIVER_LICENSE": {
        "label": "운전면허번호",
        "regex": r"[0-9]{2}-[0-9]{2}-[0-9]{6}-[0-9]{2}",
        "score": 10
    },
    "CREDIT_CARD": {
        "label": "신용카드번호",
        "regex": r"(?:4\d{3}|5[1-5]\d{2}|3[47]\d{2}|6(?:011|5\d{2}))-?\d{4}-?\d{4}-?\d{4}",
        "score": 8
    },
    "BANK_ACCOUNT": {
        "label": "계좌번호",
        "regex": r"\d{2,4}-?\d{3,6}-?\d{3,6}",
        "score": 8
    },
    "NAME": {
        "label": "이름(한글)",
        "regex": r"(?<![가-힣])[가-힣]{2,4}(?![가-힣])",
        "score": 5
    },
    "PHONE_NUMBER": {
        "label": "전화번호",
        "regex": r"(?:01[016789]-\d{3,4}-\d{4})|(?:0[2-8]\d?-\d{3,4}-\d{4})",
        "score": 5
    },
    "ADDRESS": {
        "label": "주소(상세)",
        "regex": r"(?:[가-힣0-9]+(시|도)\s*[가-힣0-9]+(구|군)|[가-힣0-9]+(동|읍|면)\s*\d+-?\d*)",
        "score": 5
    },
    "IP_ADDRESS": {
        "label": "IP 주소",
        "regex": r"\b(?:\d{1,3}\.){3}\d{1,3}\b",
        "score": 3
    },
    "MAC_ADDRESS": {
        "label": "MAC 주소",
        "regex": r"\b(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}\b",
        "score": 3
    },
    "COOKIE": {
        "label": "쿠키 / 세션 ID",
        "regex": r"(?i)\b(?:JSESSIONID|SESSIONID|SID|SESSION_TOKEN)=([A-Za-z0-9\-_]{16,})",
        "score": 3
    },
    "LOG_ENTRY": {
        "label": "접속 기록(로그)",
        "regex": r"\b(GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS)\s+\/[^\s]*\s+HTTP\/1\.[01]\b",
        "score": 3
    }
}

RISKY_EXTENSIONS = {
    'csv', 'xlsx', 'xls', 'pdf', 'txt', 'doc', 'docx', 'hwp',
    'zip', '7z', 'rar', 'tar', 'gz', 'tgz',
    'min', 'face'
}

file_id_counter = 100


def is_risky_file(filename: str) -> bool:
    """1차 탐지: 파일명/확장자로 위험 파일 판단"""
    ext = filename.lower().split('.')[-1] if '.' in filename else ''
    
    if ext in RISKY_EXTENSIONS:
        print(f"   [1차 탐지] 위험 확장자 감지: .{ext}")
        return True
    
    for type_code, pattern_info in PII_PATTERNS.items():
        if re.search(pattern_info['regex'], filename):
            print(f"   [1차 탐지] 파일명에서 {pattern_info['label']} 패턴 감지")
            return True
    
    return False


def mask_pii_value(value: str, type_code: str) -> str:
    """개인정보 마스킹 처리"""
    try:
        if type_code == "KR_RRN":
            return re.sub(r'(\d{6})-(\d)\d{6}', r'\1-\2******', value)
        elif type_code == "CREDIT_CARD":
            return re.sub(r'(\d{4})-?\d{4}-?\d{4}-?(\d{4})', r'\1-****-****-\2', value)
        elif type_code == "PHONE_NUMBER":
            return re.sub(r'(\d{2,3})-(\d{3,4})-\d{4}', r'\1-\2-****', value)
        elif type_code == "BANK_ACCOUNT":
            parts = value.split('-')
            if len(parts) >= 3:
                return f"{parts[0]}-****-{parts[-1][:2]}**"
            return "****"
        elif type_code == "MAC_ADDRESS":
            return re.sub(r'([0-9A-Fa-f]{2}[:-])([0-9A-Fa-f]{2}[:-]){3}([0-9A-Fa-f]{2})', r'\1**:**:**:\3', value)
        else:
            if len(value) > 6:
                return value[:3] + '*' * (len(value) - 6) + value[-3:]
            return '*' * len(value)
    except:
        return '***'


def detect_pii_in_text(text: str, location: str) -> List[Dict]:
    """텍스트에서 PII 탐지 및 점수화"""
    findings = []
    finding_id = 1
    
    for type_code, pattern_info in PII_PATTERNS.items():
        try:
            pattern = pattern_info['regex']
            matches = re.finditer(pattern, text)
            
            for match in matches:
                detected_value = match.group(0)
                
                findings.append({
                    "id": finding_id,
                    "type_label": pattern_info['label'],
                    "score": pattern_info['score'],
                    "location": location,
                    "value_preview": mask_pii_value(detected_value, type_code)
                })
                finding_id += 1
                
        except Exception as e:
            print(f"      [오류] 패턴 매칭 실패 ({type_code}): {e}")
            continue
    
    return findings


def extract_text_from_pdf(file_content: bytes) -> str:
    if PyPDF2 is None:
        return ""
    try:
        pdf_file = io.BytesIO(file_content)
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            extracted = page.extract_text()
            if extracted:
                text += extracted + "\n"
        return text
    except:
        return ""

def extract_text_from_docx(file_content: bytes) -> str:
    if docx is None:
        return ""
    try:
        doc = docx.Document(io.BytesIO(file_content))
        return "\n".join([para.text for para in doc.paragraphs])
    except:
        return ""

def extract_text_from_xlsx(file_content: bytes) -> str:
    if load_workbook is None:
        return ""
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except:
        return ""

def extract_text_from_csv(file_content: bytes) -> str:
    try:
        text_content = file_content.decode('utf-8', errors='ignore')
        if not text_content.strip():
            text_content = file_content.decode('cp949', errors='ignore')
        
        reader = csv.reader(io.StringIO(text_content))
        lines = []
        for row in reader:
            lines.append(" ".join(row))
        return "\n".join(lines)
    except:
        return ""

def extract_metadata_from_image(file_content: bytes) -> str:
    if Image is None or TAGS is None:
        return ""
    try:
        image = Image.open(io.BytesIO(file_content))
        exif_data = image._getexif()
        
        if not exif_data:
            return ""
        
        metadata_text = ""
        for tag_id, value in exif_data.items():
            tag_name = TAGS.get(tag_id, str(tag_id))
            metadata_text += f"{tag_name}: {str(value)} "
        
        return metadata_text
    except:
        return ""


def calculate_scores(findings: List[Dict]) -> tuple:
    """
    점수 계산 및 위험 등급 판단
    """
    
    overall_score = sum([f['score'] for f in findings]) if findings else 0
    
    if overall_score >= 8:
        risk_level = "High"
    elif overall_score >= 5:
        risk_level = "Medium"
    else:
        risk_level = "Low"
        
    return overall_score, risk_level


def analyze_file(file_id: int, filename: str, file_content: bytes, mime_type: str) -> Dict:
    """파일 분석 및 결과 도출"""
    
    print(f"\n[파일 분석] {filename}")
    
    all_findings = []
    
    try:
        # STEP 1: 파일명에서 PII 탐지
        print("   [STEP 1] 파일명 스캔")
        filename_findings = detect_pii_in_text(filename, "filename")
        all_findings.extend(filename_findings)
        if filename_findings:
            print(f"      발견: {len(filename_findings)}건")
        
        # STEP 2: 파일 내용 추출
        print("   [STEP 2] 파일 내용 추출")
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        text_content = ""
        
        if ext == 'pdf':
            text_content = extract_text_from_pdf(file_content)
            print(f"      PDF 추출 완료: {len(text_content)} 글자")
        elif ext in ['docx', 'doc']:
            text_content = extract_text_from_docx(file_content)
            print(f"      DOCX 추출 완료: {len(text_content)} 글자")
        elif ext in ['xlsx', 'xls']:
            text_content = extract_text_from_xlsx(file_content)
            print(f"      XLSX 추출 완료: {len(text_content)} 글자")
        elif ext == 'csv':
            text_content = extract_text_from_csv(file_content)
            print(f"      CSV 추출 완료: {len(text_content)} 글자")
        elif ext == 'txt':
            text_content = file_content.decode('utf-8', errors='ignore')
            print(f"      TXT 추출 완료: {len(text_content)} 글자")
        elif ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
            metadata_text = extract_metadata_from_image(file_content)
            if metadata_text:
                print("      이미지 메타데이터 추출 완료")
                metadata_findings = detect_pii_in_text(metadata_text, "metadata")
                all_findings.extend(metadata_findings)
        
        # STEP 3: 텍스트 내용에서 PII 탐지
        if text_content and text_content.strip():
            print("   [STEP 3] 내용 스캔")
            content_findings = detect_pii_in_text(text_content, "content")
            all_findings.extend(content_findings)
            if content_findings:
                print(f"      발견: {len(content_findings)}건")
        
        # STEP 4: 점수 계산 
        print("   [STEP 4] 점수 계산")
        overall_score, risk_level = calculate_scores(all_findings)
        
        print(f"   [결과] 총 {len(all_findings)}건 / 점수: {overall_score} / 등급: {risk_level}")
        
        return {
            "id": file_id,
            "filename": filename,
            "scan_status": "ok",
            "fail_reason": None,
            "overall_score": overall_score,
            "risk_level": risk_level,
            "findings": all_findings
        }
        
    except Exception as e:
        print(f"   [오류] 분석 실패: {e}")
        traceback.print_exc()
        
        return {
            "id": file_id,
            "filename": filename,
            "scan_status": "failed",
            "fail_reason": "corrupted",
            "overall_score": 0,
            "risk_level": "N/A",
            "findings": []
        }


@app.route('/api/analyze', methods=['POST'])
def analyze_files():
    """파일 분석 API"""
    global file_id_counter
    
    try:
        print("\n" + "="*60)
        print("[API 호출] 파일 분석 시작")
        print("="*60)
        
        if 'files' not in request.files:
            return jsonify({"error": "파일이 제공되지 않았습니다."}), 400
        
        files = request.files.getlist('files')
        print(f"[수신] {len(files)}개 파일")
        
        results = []
        
        for file in files:
            file_id = file_id_counter
            file_id_counter += 1
            
            filename = file.filename
            file_content = file.read()
            mime_type = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
            
            print(f"\n{'='*60}")
            print(f"[파일 ID: {file_id}] {filename}")
            print(f"{'='*60}")
            
            # 1차 탐지
            print("[1차 탐지] 파일명/확장자 검사")
            if not is_risky_file(filename):
                print("   안전한 파일 (분석 생략)")
                results.append({
                    "id": file_id,
                    "filename": filename,
                    "scan_status": "ok",
                    "fail_reason": None,
                    "overall_score": 0,
                    "risk_level": "Low",
                    "findings": []
                })
                continue
            
            # 압축 파일
            ext = filename.lower().split('.')[-1] if '.' in filename else ''
            if ext in ['zip', '7z', 'rar', 'tar', 'gz', 'tgz']:
                print("   압축 파일 (스캔 불가)")
                results.append({
                    "id": file_id,
                    "filename": filename,
                    "scan_status": "failed",
                    "fail_reason": "encrypted",
                    "overall_score": 0,
                    "risk_level": "N/A",
                    "findings": []
                })
                continue
            
            # 생체 정보 파일
            if ext in ['min', 'face']:
                print("   생체 정보 파일")
                results.append({
                    "id": file_id,
                    "filename": filename,
                    "scan_status": "ok",
                    "fail_reason": None,
                    "overall_score": 8,
                    "risk_level": "High",
                    "findings": [{
                        "id": 1,
                        "type_label": "지문 정보" if ext == 'min' else "얼굴 정보",
                        "score": 8,
                        "location": "content",
                        "value_preview": "생체 정보 파일"
                    }]
                })
                continue
            
            # 2차 탐지
            print("[2차 탐지] 파일 내용 분석")
            result = analyze_file(file_id, filename, file_content, mime_type)
            results.append(result)
        
        print("\n" + "="*60)
        print(f"[완료] 총 {len(results)}개 파일 분석 완료")
        print("="*60 + "\n")

        try:
            FASTAPI_URL = "http://localhost:8000/api/save-results"

            response = requests.post(
                FASTAPI_URL,
                json={"results": results},
                timeout=5
            )

            if response.status_code == 200:
                print(f"[연동] FastAPI 서버에 분석 결과 저장 및 가이드라인 생성 완료")
            else:
                print(f"[경고] FastAPI 서버 연동 실패 (상태 코드: {response.status_code})")
        except Exception as e:
            print(f"[오류] FastAPI 서버로 데이터 전송 중 에러 발생: {e}")
        
        return jsonify({"results": results}), 200
        
    except Exception as e:
        print(f"\n[오류] {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """서버 상태 확인"""
    return jsonify({
        "status": "healthy",
        "version": "1.0",
        "role": "Backend A - Detection Engine Only"
    }), 200


if __name__ == '__main__':
    print("\n" + "="*60)
    print("PII Detection System")
    print("="*60)
    print("Server: http://localhost:5000")
    print("API: POST http://localhost:5000/api/analyze")
    print("="*60 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=True)